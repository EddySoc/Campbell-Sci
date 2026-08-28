from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference
from openpyxl.chart.data_source import AxDataSource, NumRef
from openpyxl.utils import get_column_letter

# Create workbook with data
wb = Workbook()
ws = wb.active
ws.title = "Data"

# Add simple data: 2 columns with numbers
ws['A1'] = 'OldSensor'
ws['B1'] = 'NewSensor'

for i in range(2, 12):
    ws[f'A{i}'] = i - 1  # 1, 2, 3, ..., 10
    ws[f'B{i}'] = (i - 1) * 1.1  # 1.1, 2.2, 3.3, ..., 11

print("Data in worksheet:")
for row in ws.iter_rows(min_row=1, max_row=11, min_col=1, max_col=2, values_only=True):
    print(row)

# Create scatter chart
chart = ScatterChart()
chart.title = "Sensor Comparison"
chart.x_axis.title = 'New Sensor'
chart.y_axis.title = 'Old Sensor'

# Add data: Y values (OldSensor) - WITH header row
y_data = Reference(ws, min_col=1, min_row=1, max_row=11)
chart.add_data(y_data, titles_from_data=True)

# Set X values (NewSensor) - data rows only, no header
x_data = Reference(ws, min_col=2, min_row=2, max_row=11)
# Convert to AxDataSource (required by openpyxl)
col_letter = get_column_letter(x_data.min_col)
formula = f"Data!${col_letter}${x_data.min_row}:${col_letter}${x_data.max_row}"
numref = NumRef(f=formula)
ax_data_source = AxDataSource(numRef=numref)
chart.series[0].xVal = ax_data_source

print(f"\nChart series count: {len(chart.series)}")
print(f"Series 0 val: {chart.series[0].val}")
print(f"Series 0 xVal: {chart.series[0].xVal}")
print(f"Series 0 title: {chart.series[0].title}")

# Add to sheet
ws2 = wb.create_sheet("Chart")
ws2.add_chart(chart)

# Save
output_path = "test_simple_scatter_output.xlsx"
wb.save(output_path)
print(f"\nSaved to {output_path}")
