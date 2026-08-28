"""Python-versie van de VBA-macro `Builder` (zie GraphBuilder.txt).

Bouwt één grafiekblad met meerdere reeksen uit een gegevensblad, met
ondersteuning voor een primaire en secundaire Y-as - net als de originele
VBA-routine. Dit vervangt de "fallback" (één chartsheet per kolom) door één
gecombineerde grafiek met de kolommen die je zelf kiest.

Gebruik:
    from openpyxl import load_workbook
    from campbell_sci.graph_builder import build_chart_sheet

    wb = load_workbook("resultaat.xlsx")
    build_chart_sheet(
        wb,
        chart_name="Temperatuur",
        data_sheet="Data",
        chart_type="line",
        columns="A B c",   # A = tijd/categorie, B = primaire as, c = secundaire as
        num_format="0.0",
        tab_color="1F4E78",
    )
    wb.save("resultaat.xlsx")
"""

from __future__ import annotations

import datetime
import math
import re
from io import BytesIO
from pathlib import Path
from typing import Any, Sequence, cast

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.data_source import AxDataSource, NumRef, NumDataSource
from openpyxl.chart.series import Series, SeriesLabel
from openpyxl.chart.series_factory import SeriesFactory
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.chartsheet.properties import ChartsheetProperties
from openpyxl.styles.colors import Color
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D

# Brede (16:9) afmeting voor het grafiekblad, zodat een breedbeeldscherm zoveel
# mogelijk gevuld wordt in plaats van de portret-achtige printpagina-verhouding
# die Excel anders als standaard aanhoudt.
CHART_SHEET_SIZE = (24_384_000, 13_716_000)  # EMU, 64 x 36 cm

# Interne kolomtitel voor de verborgen "1 label per dag"-hulpkolom.
DAY_LABEL_MARKER = "__DagLabel__"

# Kleuren die na elkaar toegekend worden aan reeksen (equivalent van
# MFuncties.LijnKleur / OpVulKleur in de VBA-versie).
LINE_COLORS = ["1F4E78", "C00000", "548235", "BF8F00", "7030A0", "2E75B6", "833C00"]
FILL_COLORS = ["9DC3E6", "F4B183", "A9D18E", "FFD966", "B4A7D6", "8EA9DB", "D9D2E9"]

CHART_TYPES = {"line": LineChart, "column": BarChart, "scatter": ScatterChart}
CONFIG_DIR = Path(__file__).with_name("graph_configs")
_CELL_RANGE_RE = re.compile(r"(\$?[A-Z]+\$?)\d+(?::(\$?[A-Z]+\$?))?\d+$")


def _next_color(palette: Sequence[str], index: int) -> str:
    return palette[index % len(palette)]


def _reference_formula(ref: Reference) -> str:
    """Build a formula string from a Reference object for use with NumRef."""
    col_letter = get_column_letter(ref.min_col)
    # ref.sheetname is already quoted, so don't add extra quotes
    sheetname = ref.sheetname if ref.sheetname.startswith("'") else f"'{ref.sheetname}'"
    return f"{sheetname}!${col_letter}${ref.min_row}:${col_letter}${ref.max_row}"


def _chart_objects(workbook: Workbook, chart_name: str):
    if chart_name not in workbook.sheetnames:
        raise ValueError(f"Grafiekblad '{chart_name}' niet gevonden.")
    charts = getattr(workbook[chart_name], "_charts", None)
    if charts is None:
        raise ValueError(f"'{chart_name}' is geen grafiekblad.")
    return charts


def _series_values(ws, series, start_row: int | None = None, end_row: int | None = None) -> list[float]:
    reference = getattr(getattr(series, "val", None), "numRef", None)
    formula = getattr(reference, "f", "") if reference else ""
    match = re.search(r"!\$?[A-Z]+\$?(\d+):\$?[A-Z]+\$?(\d+)$", formula or "")
    first = start_row or (int(match.group(1)) if match else 2)
    last = end_row or (int(match.group(2)) if match else ws.max_row)
    column_match = re.search(r"!\$?([A-Z]+)", formula or "")
    if not column_match:
        return []
    column = column_index_from_string(column_match.group(1))
    values = []
    for row in range(first, min(last, ws.max_row) + 1):
        value = ws.cell(row=row, column=column).value
        if isinstance(value, (int, float)) and math.isfinite(value):
            values.append(float(value))
    return values


def _set_formula_rows(formula: str | None, start_row: int, end_row: int) -> str | None:
    if not formula:
        return formula
    return re.sub(
        r"(\$?[A-Z]+\$?)\d+:(\$?[A-Z]+\$?)\d+$",
        rf"\g<1>{start_row}:\g<2>{end_row}",
        formula,
    )


def _set_series_rows(series, start_row: int, end_row: int) -> None:
    for attribute in ("val", "xVal", "yVal"):
        container = getattr(series, attribute, None)
        if container is None:
            continue
        for reference_name in ("numRef", "strRef"):
            reference = getattr(container, reference_name, None)
            if reference is not None:
                reference.f = _set_formula_rows(reference.f, start_row, end_row)


def zoom_chart(workbook: Workbook, chart_name: str, start_row: int, end_row: int) -> None:
    """Show only the selected data rows, equivalent to VBA ``ZoomIn``."""
    if start_row < 2 or end_row < start_row:
        raise ValueError("start_row moet >= 2 zijn en end_row moet groter zijn.")
    for chart in _chart_objects(workbook, chart_name):
        for series in chart.series:
            _set_series_rows(series, start_row, end_row)


def scroll_chart(workbook: Workbook, chart_name: str, direction: str, step: int) -> tuple[int, int]:
    """Move the current line-chart window left or right and return its rows."""
    charts = _chart_objects(workbook, chart_name)
    series = charts[0].series[0]
    reference = getattr(getattr(series, "val", None), "numRef", None)
    formula = getattr(reference, "f", "") if reference else ""
    match = re.search(r"\$(\d+):\$?[A-Z]+\$(\d+)$", formula or "")
    if not match:
        raise ValueError(f"Geen zoombereik gevonden voor grafiek '{chart_name}'.")
    start_row, end_row = map(int, match.groups())
    delta = step if direction.lower() in {"right", "rechts"} else -step
    new_start = max(2, start_row + delta)
    new_end = end_row + (new_start - start_row)
    last_row = workbook["Data"].max_row
    if new_end > last_row:
        new_end = last_row
        new_start = max(2, new_end - (end_row - start_row))
    zoom_chart(workbook, chart_name, new_start, new_end)
    return new_start, new_end


def rescale_chart(workbook: Workbook, chart_name: str, mode: str = "auto", padding: float = 0.1) -> None:
    """Rescale chart axes like VBA ``Rescale``.

    ``auto`` scales each axis independently, ``split`` adds room below the
    primary range and above the secondary range, and ``same_zero`` gives both
    axes the same combined range.
    """
    if mode not in {"auto", "split", "same_zero"}:
        raise ValueError("mode moet 'auto', 'split' of 'same_zero' zijn.")
    charts = _chart_objects(workbook, chart_name)
    data_sheet = next(sheet for sheet in workbook.worksheets if sheet.title == "Data")
    chart_ranges = [[value for series in chart.series for value in _series_values(data_sheet, series)] for chart in charts]
    if not any(chart_ranges):
        return
    if mode == "same_zero":
        values = [value for values in chart_ranges for value in values]
        minimum, maximum = min(values), max(values)
        for chart in charts:
            chart.y_axis.scaling.min = minimum
            chart.y_axis.scaling.max = maximum
        return
    for index, values in enumerate(chart_ranges):
        if not values:
            continue
        minimum, maximum = min(values), max(values)
        span = maximum - minimum or max(abs(maximum), 1.0)
        if mode == "split":
            if index == 0:
                minimum -= span
            else:
                maximum += span
        else:
            minimum -= span * padding
            maximum += span * padding
        charts[index].y_axis.scaling.min = minimum
        charts[index].y_axis.scaling.max = maximum


def load_chart_config(logger_name: str, config_dir: str | Path = CONFIG_DIR) -> list[dict]:
    """Read one chart definition per line from a logger .logdef file."""
    path = Path(config_dir) / f"{logger_name.upper()}.logdef"
    if not path.is_file():
        return []
    charts = []
    for line_number, raw_line in enumerate(path.read_text(encoding="utf-8").splitlines(), 1):
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        fields = [field.strip() for field in line.split(";")]
        if len(fields) != 7:
            raise ValueError(f"Ongeldige chart-configuratie in {path}:{line_number}")
        name, chart_type, columns, minimum, maximum, num_format, tab_color = fields
        charts.append({
            "chart_name": name,
            "chart_type": chart_type,
            "columns": columns,
            "minimum": float(minimum) if minimum else None,
            "maximum": float(maximum) if maximum else None,
            "num_format": num_format or "0.0",
            "tab_color": tab_color or None,
        })
    return charts


def _columns_from_headers(ws, column_spec: str) -> str:
    headers = {str(cell.value).strip().upper(): cell.column for cell in ws[1]}
    tokens = []
    for configured in column_spec.split(","):
        configured = configured.strip()
        secondary = configured.lower().endswith("@secondary")
        header = configured[:-10].strip() if secondary else configured
        if header.upper() not in headers:
            raise ValueError(f"Kolom '{header}' niet gevonden op werkblad '{ws.title}'.")
        letter = get_column_letter(headers[header.upper()])
        tokens.append(letter.lower() if secondary else letter)
    return " ".join(tokens)


def build_configured_charts(workbook: Workbook, logger_name: str, data_sheet: str = "Data", first_data_row: int = 2, config_dir: str | Path = CONFIG_DIR) -> int:
    """Build all charts declared in the logger's text configuration."""
    ws = workbook[data_sheet]
    count = 0
    for chart in load_chart_config(logger_name, config_dir):
        chart["columns"] = _columns_from_headers(ws, chart["columns"])
        build_chart_sheet(workbook, data_sheet=data_sheet, first_data_row=first_data_row, **chart)
        count += 1
    return count


def build_fallback_charts(workbook: Workbook, data_sheet: str = "Data", first_data_row: int = 2) -> int:
    """Create one consistently styled chart per measurement column."""
    ws = workbook[data_sheet]
    count = 0
    for col_idx in range(2, ws.max_column + 1):
        header = str(ws.cell(row=1, column=col_idx).value or "").strip()
        if not header or header == DAY_LABEL_MARKER:
            continue
        build_chart_sheet(workbook, header[:31], data_sheet, "line", f"A {get_column_letter(col_idx)}", num_format="0.0", first_data_row=first_data_row)
        count += 1
    return count


def _make_chart(chart_type: str):
    chart = CHART_TYPES.get(chart_type, LineChart)()
    if chart_type == "column":
        chart.type = "col"
        chart.grouping = "clustered"
    return chart


def _day_label_column(ws, time_col_idx: int, first_data_row: int) -> int:
    """Geef de kolom met 1 zichtbaar label per dag (enkel bij 00u00) terug.

    Wordt eenmalig aangemaakt per werkblad en hergebruikt door alle grafieken
    die er kolommen uit gebruiken, zodat we hem niet telkens dupliceren.
    """
    for cell in ws[1]:
        if cell.value == DAY_LABEL_MARKER:
            return cell.column

    label_col = ws.max_column + 1
    ws.cell(row=1, column=label_col, value=DAY_LABEL_MARKER)
    for row in range(first_data_row, ws.max_row + 1):
        timestamp = ws.cell(row=row, column=time_col_idx).value
        is_midnight = isinstance(timestamp, datetime.datetime) and timestamp.time() == datetime.time(0, 0)
        label_cell = ws.cell(row=row, column=label_col)
        if is_midnight:
            label_cell.value = timestamp
            label_cell.number_format = "dd/mm/yyyy"
        else:
            label_cell.value = None
    return label_col


def _style_value_axis(
    chart,
    num_format: str,
    minimum,
    maximum,
    dotted_gridlines: bool,
    show_category_axis: bool = True,
):
    chart.y_axis.numFmt = num_format or "0.0"
    chart.y_axis.scaling.min = minimum
    chart.y_axis.scaling.max = maximum
    chart.y_axis.delete = False
    # Combineren van een primaire + secundaire grafiek geeft een tweede,
    # ongebruikte categorie-as die anders als een lege/foute as verschijnt.
    chart.x_axis.delete = not show_category_axis
    # openpyxl zet axPos standaard op "l" (links) i.p.v. "b" (onderaan); bij
    # meerdere reeksen corrigeert Excel dit stilzwijgend, bij één reeks niet -
    # vandaar de rommelige, verticaal opgestapelde labels.
    chart.x_axis.axPos = "b"
    if dotted_gridlines:
        chart.y_axis.majorGridlines = ChartLines(
            spPr=GraphicalProperties(ln=LineProperties(prstDash="sysDot"))
        )


def _add_scatter_reference_line(chart, min_val: float, max_val: float):
    """Add a 45° reference line (diagonal) to a scatter chart to show perfect calibration.
    
    The reference line connects (min_val, min_val) to (max_val, max_val).
    Since we can't easily add raw trendlines in openpyxl, we add a helper line
    by manually modifying the chart XML after it's added to the workbook.
    """
    # For now, we'll just set up the axes scaling correctly.
    # The reference line will need to be added via XML manipulation after wb.save()
    # or via a post-processing step. This is a known limitation of openpyxl.
    pass


def build_chart_sheet(
    workbook: Workbook,
    chart_name: str,
    data_sheet: str,
    chart_type: str,
    columns: str,
    minimum: float | None = None,
    maximum: float | None = None,
    num_format: str = "0.0",
    date_format: str | None = None,
    tab_color: str | None = None,
    first_data_row: int = 2,
):
    """Bouw één grafiekblad met meerdere reeksen, net als de VBA `Builder`-macro.

    - `columns`: kolomletters gescheiden door spaties, bv. `"A B c D"`.
      De eerste letter is de categorie/tijd-as (of de X-as bij een scatterplot).
      Voor de overige kolommen geldt dezelfde conventie als in VBA: een
      kolomletter in KLEINE letters (bv. `"c"` i.p.v. `"C"`) plaatst die reeks
      op de SECUNDAIRE Y-as, hoofdletters blijven op de primaire Y-as.
    - `minimum` / `maximum`: vaste schaal voor de primaire Y-as (None = automatisch).
    - `first_data_row`: eerste rij met echte meetwaarden (rij 1 = headers).
      Gebruik dit om metadata-rijen (eenheden, "Avg"/"Max" ...) uit de
      Campbell-export over te slaan; anders overspoelen die de categorie-as.
    """
    ws = workbook[data_sheet]
    last_row = ws.max_row
    is_scatter = chart_type == "scatter"

    tokens = columns.split()
    if is_scatter:
        if len(tokens) != 2:
            raise ValueError("Een scattergrafiek vereist exact twee kolommen: Y en X.")
        y_idx = column_index_from_string(tokens[0].upper())
        x_idx = column_index_from_string(tokens[1].upper())
        y_title = str(ws.cell(row=1, column=y_idx).value or tokens[0])
        x_title = str(ws.cell(row=1, column=x_idx).value or tokens[1])
    elif len(tokens) < 2:
        raise ValueError("Geef minstens een categoriekolom en één datakolom op.")

    if not is_scatter:
        cat_idx = column_index_from_string(tokens[0].upper())
        # For non-scatter: Use day-label column to show only 1 label per day (at 00:00)
        # This avoids cluttering the x-axis with too many timestamps
        label_col = _day_label_column(ws, cat_idx, first_data_row)
        categories = Reference(ws, min_col=label_col, min_row=first_data_row, max_row=last_row)

    primary = _make_chart(chart_type)
    secondary = _make_chart(chart_type)
    line_i = fill_i = 0
    has_secondary = False

    scatter_series = None
    series_tokens = (tokens[0],) if is_scatter else tokens[1:]
    for token in series_tokens:
        is_primary = token == token.upper()
        col_idx = column_index_from_string(token.upper())
        title = ws.cell(row=1, column=col_idx).value
        
        target = primary if is_primary else secondary
        if is_scatter:
            y_ref = Reference(ws, min_col=y_idx, min_row=first_data_row, max_row=last_row)
            x_ref = Reference(ws, min_col=x_idx, min_row=first_data_row, max_row=last_row)
            series = SeriesFactory(
                y_ref,
                xvalues=x_ref,
                title=f"{y_title} vs {x_title}",
            )
            target.series.append(series)
        else:
             # Include the header so openpyxl can use it as the series title.
             data = Reference(ws, min_col=col_idx, min_row=1, max_row=last_row)
             target.add_data(data, titles_from_data=True)
             series = target.series[-1]

        if chart_type == "column":
            series.graphicalProperties.solidFill = _next_color(FILL_COLORS, fill_i)
            series.graphicalProperties.line.solidFill = "000000"
            fill_i += 1
        elif chart_type != "scatter":  # Don't smooth scatter charts
            series.graphicalProperties.line.solidFill = _next_color(LINE_COLORS, line_i)
            series.graphicalProperties.line.width = 20000 if is_primary else 28000
            series.smooth = True
            line_i += 1
        else:  # scatter chart
            series.graphicalProperties.line.solidFill = _next_color(LINE_COLORS, line_i)
            series.graphicalProperties.line.width = 20000
            line_i += 1

        has_secondary = has_secondary or not is_primary
        if is_scatter:
            scatter_series = series

    if not is_scatter:
        primary.set_categories(categories)

    if is_scatter:
        # For scatter charts: auto-scale both axes (let Excel determine from data)
        # Don't set explicit minimum/maximum - Excel will auto-scale based on actual data
        primary.x_axis.scaling.min = None  # Auto
        primary.x_axis.scaling.max = None  # Auto
        primary.y_axis.scaling.min = None  # Auto
        primary.y_axis.scaling.max = None  # Auto
        primary.x_axis.delete = False  # Show X-axis
        primary.y_axis.delete = False  # Show Y-axis
        primary.y_axis.numFmt = num_format or "0.0"
        primary.title = f"{y_title} vs {x_title}"
    else:
        # For non-scatter charts: use configured scaling
        _style_value_axis(primary, num_format, minimum, maximum, dotted_gridlines=True)
    
    if len(primary.series) > 1 or has_secondary:
        primary.legend.position = "t"
    else:
        primary.legend = None

    if not is_scatter and date_format:
        primary.x_axis.number_format = date_format

    if has_secondary:
        if not is_scatter:
            secondary.set_categories(categories)
        secondary.y_axis.axId = 200
        _style_value_axis(secondary, num_format, None, None, dotted_gridlines=False, show_category_axis=False)
        primary.y_axis.crosses = "min"
        primary += secondary

    # Vervang een bestaand grafiekblad met dezelfde naam.
    if chart_name in workbook.sheetnames:
        del workbook[chart_name]

    chart_sheet = workbook.create_chartsheet(title=chart_name)
    if tab_color:
        chart_sheet.sheetPr = ChartsheetProperties(tabColor=Color(rgb=f"FF{tab_color.upper()}"))
    chart_sheet.add_chart(primary)
    cx, cy = CHART_SHEET_SIZE
    primary.anchor = AbsoluteAnchor(pos=XDRPoint2D(0, 0), ext=XDRPositiveSize2D(cx, cy))

    return chart_sheet
