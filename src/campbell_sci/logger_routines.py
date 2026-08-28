"""Python-vertaling van `MDataLogger.txt` / `MCommon.txt` (VBA).

Elke functie hier komt overeen met één specifieke datalogger-routine uit de
VBA-module `MDataLogger`, geselecteerd op basis van de loggernaam die uit de
bestandsnaam gehaald wordt (zie `campbell_sci.parser.parse_filename`).

Voorlopig is enkel de METEO-routine (logger "BM", bv. `bm_2011_01_18_..._.dat`)
overgezet. De overige routines (TDR_PA, WX_Station, IMMISSIE, ICOS_*) volgen op
dezelfde manier zodra we ze één voor één aanpakken.

Let op - bewuste vereenvoudiging t.o.v. de VBA:
De originele VBA verplaatst/verwijdert eerst enkele kolommen (`MKolom.Verwijderen`,
`MKolom.Verplaatsen`) en berekent een windrichtingslabel (`MKolom.Windrichting`)
voordat de grafieken gebouwd worden. Die hulpmodules (`MKolom`, `MFuncties`) zijn
niet meegeleverd, dus kunnen we de exacte kolomherschikking niet 1-op-1 nabouwen.
Om dat te omzeilen zoeken de grafieken hier kolommen op via de HEADERNAAM
(bv. "Wind_Dir_WVT") in plaats van een vaste letterpositie ("AG") - dat blijft
correct ongeacht kolomvolgorde. De "Decagon"-grafiek uit de VBA is nog niet
overgezet omdat die verwijst naar kolommen die pas na de (onbekende)
kolomherschikking hun definitieve positie krijgen.
"""

from __future__ import annotations

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from campbell_sci.graph_builder import build_chart_sheet

# VBA Tabkleur (ColorIndex) -> hex, enkel de kleuren die MDataLogger.METEO gebruikt.
TAB_COLORS = {
    3: "FF0000",  # Rood
    5: "0000FF",  # Blauw
    6: "FFFF00",  # Geel
    8: "00FFFF",  # Turkoise
}


def _letter(ws: Worksheet, header: str, secondary: bool = False) -> str:
    """Zoek de kolomletter op basis van de headernaam (rij 1).

    Kleine letters = secundaire Y-as, exact dezelfde conventie als
    `build_chart_sheet`/de originele VBA (`kol(XX) = UCase(kol(XX))`).
    """
    for cell in ws[1]:
        if str(cell.value).strip().upper() == header.strip().upper():
            letter = get_column_letter(cell.column)
            return letter.lower() if secondary else letter
    raise ValueError(f"Kolom '{header}' niet gevonden op werkblad '{ws.title}'.")


def mask_bad_values(df):
    """Python-versie van `MCommon.Maskeren`: sentinelwaarden vervangen door N/A (NaN)."""
    bad_values = ["6999", "-6999", "7999", "-7999", "NAN", "INF"]
    return df.replace(bad_values, None)


def add_leaf_temperature(df, ir_out_column: str = "IR_R_R_Avg", label: str = "Leaf T."):
    """Python-versie van de VBA-formule `=(S1/0.0000000567)^(1/4)-273.16`."""
    df = df.copy()
    ir_out = pd.to_numeric(df[ir_out_column], errors="coerce")
    df[label] = (ir_out / 0.0000000567) ** 0.25 - 273.16
    return df


def process_meteo(workbook, data_sheet: str = "Data", first_data_row: int = 2):
    """Build BM charts from the external logger configuration."""
    from campbell_sci.graph_builder import build_configured_charts
    return build_configured_charts(workbook, "BM", data_sheet, first_data_row)
