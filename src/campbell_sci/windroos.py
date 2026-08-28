"""Windroosgrafiek toevoegen aan een Excel-bestand op basis van één kolom.

Python-vertaling van de VBA-macro `WindRoos`. In plaats van te werken met een
vast Excel-sjabloon (het "WNDRs"-werkblad in de VBA-versie), wordt de
windroos hier opnieuw opgebouwd met matplotlib (polaire staafgrafiek) en als
afbeelding in een nieuw werkblad geplaatst.

Gebruik:
    from campbell_sci.windroos import add_wind_rose
    add_wind_rose("resultaat.xlsx", column="WindDir")
"""

from __future__ import annotations

import os
import re
import tempfile
from pathlib import Path
from typing import Sequence

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import column_index_from_string

COMPASS_LABELS = [
    "N", "NNE", "NE", "ENE", "E", "ESE", "SE", "SSE",
    "S", "SSW", "SW", "WSW", "W", "WNW", "NW", "NNW",
]

# VBA kleurindex 8 komt overeen met turquoise (RGB 00FFFF).
TAB_KLEUR_TURQUOISE = "00FFFF"


def _resolve_column(ws, column: str | int) -> int:
    """Vind de kolomindex op basis van een kolomletter, headernaam of index."""
    if isinstance(column, int):
        return column
    if re.fullmatch(r"[A-Za-z]+", column):
        return column_index_from_string(column.upper())

    for cell in ws[1]:
        if str(cell.value).strip().upper() == column.strip().upper():
            return cell.column

    raise ValueError(f"Kolom '{column}' niet gevonden op werkblad '{ws.title}'.")


def _read_directions(ws, col_idx: int) -> list[float]:
    """Lees alle numerieke windrichtingen (graden) uit één kolom."""
    directions: list[float] = []
    for (value,) in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
        if value is None:
            continue
        try:
            directions.append(float(str(value).replace(",", ".")))
        except ValueError:
            continue

    if not directions:
        raise ValueError("Geen numerieke windrichtingsgegevens gevonden in de opgegeven kolom.")

    return directions


def _bin_directions(directions: Sequence[float], bins: int = 16) -> list[float]:
    """Verdeel windrichtingen over kompasrichtingen en geef percentages terug."""
    bin_width = 360 / bins
    counts = [0] * bins
    for deg in directions:
        deg = deg % 360
        idx = int((deg + bin_width / 2) // bin_width) % bins
        counts[idx] += 1

    total = sum(counts)
    return [c / total * 100 for c in counts]


def _plot_wind_rose(percentages: Sequence[float], bins: int, title: str, image_path: Path) -> None:
    bin_width = 360 / bins
    theta = np.deg2rad([i * bin_width for i in range(bins)])

    fig = plt.figure(figsize=(6, 6))
    ax = fig.add_subplot(111, projection="polar")
    ax.set_theta_zero_location("N")
    ax.set_theta_direction(-1)
    ax.bar(theta, percentages, width=np.deg2rad(bin_width * 0.9), color="#1F4E78", edgecolor="white")
    ax.set_xticks(theta)
    ax.set_xticklabels(COMPASS_LABELS[:bins])
    ax.set_title(title)

    fig.savefig(image_path, dpi=150, bbox_inches="tight")
    plt.close(fig)


def add_wind_rose_to_workbook(
    workbook,
    column: str | int,
    source_sheet: str | int = 0,
    output_sheet: str = "Wind F.",
    bins: int = 16,
    image_dir: str | Path | None = None,
) -> Path:
    """Voeg een windroos-werkblad toe aan een al geopend (in-memory) workbook.

    Gebruik dit wanneer je nog meer wijzigingen aan `workbook` aanbrengt vóór
    het opslaan (zoals in `logger_routines.process_meteo`). Roep zelf
    `workbook.save(...)` aan nadat deze functie is teruggekeerd.
    """
    ws = workbook[source_sheet] if isinstance(source_sheet, str) else workbook.worksheets[source_sheet]

    col_idx = _resolve_column(ws, column)
    directions = _read_directions(ws, col_idx)
    percentages = _bin_directions(directions, bins=bins)

    image_dir = Path(image_dir) if image_dir else Path(tempfile.gettempdir())
    image_path = image_dir / f"windroos_{os.getpid()}_{id(workbook)}.png"
    _plot_wind_rose(percentages, bins, title=f"Windroos - {ws.cell(row=1, column=col_idx).value}", image_path=image_path)

    if output_sheet in workbook.sheetnames:
        del workbook[output_sheet]
    wind_ws = workbook.create_sheet(output_sheet, 0)
    wind_ws.sheet_properties.tabColor = TAB_KLEUR_TURQUOISE
    wind_ws.add_image(XLImage(str(image_path)), "B3")

    return image_path


def add_wind_rose(
    excel_path: str | Path,
    column: str | int,
    source_sheet: str | int = 0,
    output_sheet: str = "Wind F.",
    bins: int = 16,
) -> Path:
    """Voeg een windroosgrafiek toe aan `excel_path` (opent en bewaart het bestand zelf).

    - `column`: kolomletter (bv. "C"), headernaam (bv. "WindDir") of 1-based index.
    - `source_sheet`: naam of index van het werkblad met de brongegevens.
    - `output_sheet`: naam van het nieuwe werkblad met de windroos.
    """
    excel_path = Path(excel_path)
    wb = load_workbook(excel_path)

    image_path = add_wind_rose_to_workbook(wb, column, source_sheet, output_sheet, bins)
    wb.save(excel_path)
    _delete_quietly(image_path)
    print(f"Windroos toegevoegd aan werkblad '{output_sheet}' in {excel_path}")
    return excel_path


def _delete_quietly(path: str | Path) -> None:
    try:
        Path(path).unlink(missing_ok=True)
    except Exception:
        pass


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 3:
        print("Gebruik: python -m campbell_sci.windroos <excel_bestand> <kolom>")
        raise SystemExit(1)

    add_wind_rose(sys.argv[1], sys.argv[2])
