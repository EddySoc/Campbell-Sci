from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference
from openpyxl.chart.series import Series
from openpyxl.chart.data_source import AxDataSource, NumRef, NumDataSource
from openpyxl.utils import get_column_letter

def _reference_formula(ref: Reference) -> str:
    """Build a formula string from a Reference object for use with NumRef."""
    col_letter = get_column_letter(ref.min_col)
    # ref.sheetname is already quoted, so don't add extra quotes
    sheetname = ref.sheetname if ref.sheetname.startswith("'") else f"'{ref.sheetname}'"
    return f"{sheetname}!${col_letter}${ref.min_row}:${col_letter}${ref.max_row}"

wb = Workbook()
ws = wb.active

# Add data
ws['A1'] = 'X'
ws['B1'] = 'Y1'
for i in range(2, 10):
    ws[f'A{i}'] = i
    ws[f'B{i}'] = i * 2

# Create scatter chart the same way as in graph_builder.py
chart = ScatterChart()

# Create Reference objects
cat_idx = 1
data_col = 2
first_data_row = 2
last_row = 9

categories = Reference(ws, min_col=cat_idx, min_row=first_data_row, max_row=last_row)
data = Reference(ws, min_col=data_col, min_row=first_data_row, max_row=last_row)

# Manually create Series with both val and xVal
ser = Series()

# Set val (Y-axis data)
data_formula = _reference_formula(data)
print(f"data_formula: {data_formula}")
ser.val = NumDataSource(numRef=NumRef(f=data_formula))
print(f"After setting val: {ser.val}")

# Set xVal (X-axis data)
xval_formula = _reference_formula(categories)
print(f"xval_formula: {xval_formula}")
ser.xVal = AxDataSource(numRef=NumRef(f=xval_formula))
print(f"After setting xVal: {ser.xVal}")

# Set title
from openpyxl.chart.series import SeriesLabel
ser.title = SeriesLabel(v="TestData")

# Add to chart
chart.series.append(ser)

print(f"\nSeries in chart: {len(chart.series)}")
print(f"Series[0].val: {chart.series[0].val}")
print(f"Series[0].xVal: {chart.series[0].xVal}")

# Create chart sheet and save
chart_sheet = wb.create_chartsheet(title='TestChart')
chart_sheet.add_chart(chart)
wb.save('test_series_manual.xlsx')

print("Saved to test_series_manual.xlsx")

# Now verify by reading the XML
import zipfile
with zipfile.ZipFile('test_series_manual.xlsx', 'r') as z:
    content = z.read('xl/charts/chart1.xml').decode('utf-8')
    if '<ser>' in content:
        start = content.find('<ser>')
        end = content.find('</ser>') + len('</ser>')
        ser_xml = content[start:end]
        
        print("\nGenerated XML:")
        print(f"Has val: {'<val>' in ser_xml}")
        print(f"Has xVal: {'<xVal>' in ser_xml}")
        
        if '<val>' in ser_xml:
            val_start = ser_xml.find('<val>')
            val_end = ser_xml.find('</val>') + len('</val>')
            print(f"val: {ser_xml[val_start:val_end]}")
        
        if '<xVal>' in ser_xml:
            xval_start = ser_xml.find('<xVal>')
            xval_end = ser_xml.find('</xVal>') + len('</xVal>')
            print(f"xVal: {ser_xml[xval_start:xval_end]}")
