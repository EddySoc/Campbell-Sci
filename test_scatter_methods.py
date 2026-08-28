from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference
from openpyxl.chart.series import Series
from openpyxl.chart.data_source import AxDataSource, NumRef, NumDataSource
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active

# Add data: X values in column A, Y values in columns B and C
ws['A1'] = 'X'
ws['B1'] = 'Y1'
ws['C1'] = 'Y2'
for i in range(2, 10):
    ws[f'A{i}'] = i
    ws[f'B{i}'] = i * 2
    ws[f'C{i}'] = i * 3

# Create scatter chart
chart = ScatterChart()

# Try METHOD 1: add_data()
print("Method 1: Using add_data()")
ref_y1 = Reference(ws, min_col=2, min_row=2, max_row=9)  # WITHOUT header
chart.add_data(ref_y1)
print(f"  Series[0].val: {chart.series[0].val}")
print(f"  Series[0].xVal: {chart.series[0].xVal}")

# Now set xVal
ref_x = Reference(ws, min_col=1, min_row=2, max_row=9)
col_letter = get_column_letter(ref_x.min_col)
sheetname = ref_x.sheetname if ref_x.sheetname.startswith("'") else f"'{ref_x.sheetname}'"
formula = f"{sheetname}!${col_letter}${ref_x.min_row}:${col_letter}${ref_x.max_row}"
numref = NumRef(f=formula)
chart.series[0].xVal = AxDataSource(numRef=numref)

print(f"  After setting xVal:")
print(f"    Series[0].val: {chart.series[0].val}")
print(f"    Series[0].xVal: {chart.series[0].xVal}")

# Save and check XML
chart_sheet = wb.create_chartsheet(title='TestChart')
chart_sheet.add_chart(chart)
wb.save('test_scatter_method1.xlsx')
print("  Saved to test_scatter_method1.xlsx")

# Now try METHOD 2: Manually creating Series
print("\nMethod 2: Using Series with NumDataSource")
wb2 = Workbook()
ws2 = wb2.active
ws2['A1'] = 'X'
ws2['B1'] = 'Y1'
for i in range(2, 10):
    ws2[f'A{i}'] = i
    ws2[f'B{i}'] = i * 2

chart2 = ScatterChart()

# Create Series with val and xVal
ser = Series()

# Set val (Y values)
ser.val = NumDataSource(numRef=NumRef(f="'Sheet'!$B$2:$B$9"))

# Set xVal (X values)  
ser.xVal = AxDataSource(numRef=NumRef(f="'Sheet'!$A$2:$A$9"))

ser.title = 'Y1'
chart2.series.append(ser)

chart_sheet2 = wb2.create_chartsheet(title='TestChart')
chart_sheet2.add_chart(chart2)
wb2.save('test_scatter_method2.xlsx')
print("  Saved to test_scatter_method2.xlsx")
print(f"  Series[0].val: {chart2.series[0].val}")
print(f"  Series[0].xVal: {chart2.series[0].xVal}")
