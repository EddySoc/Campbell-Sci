from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.data_source import AxDataSource, NumRef
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active

# Add data: X values in column A, Y values in columns B and C
ws['A1'] = 'X'
ws['B1'] = 'Y1'
ws['C1'] = 'Y2'
for i in range(2, 10):
    ws[f'A{i}'] = i
    ws[f'B{i}'] = i * 2
    ws[f'C{i}'] = i * 3

# Create scatter chart
chart = ScatterChart()

# Create references - WITHOUT header for data
ref_x = Reference(ws, min_col=1, min_row=2, max_row=9)
ref_y1 = Reference(ws, min_col=2, min_row=2, max_row=9)  # data only
ref_y2 = Reference(ws, min_col=3, min_row=2, max_row=9)  # data only

print("Method: Correct - data references without headers")
chart.add_data(ref_y1)
chart.add_data(ref_y2)

# Set xVal for both series
col_letter = get_column_letter(ref_x.min_col)
sheetname = ref_x.sheetname if ref_x.sheetname.startswith("'") else f"'{ref_x.sheetname}'"
formula = f"{sheetname}!${col_letter}${ref_x.min_row}:${col_letter}${ref_x.max_row}"
print(f"Formula for X values: {formula}")

for i in range(len(chart.series)):
    numref = NumRef(f=formula)
    chart.series[i].xVal = AxDataSource(numRef=numref)
    if i == 0:
        chart.series[i].title = SeriesLabel(v="Y1")
    else:
        chart.series[i].title = SeriesLabel(v="Y2")

print(f"Series 0 val: {chart.series[0].val}")
print(f"Series 0 xVal: {chart.series[0].xVal}")
print(f"Series 1 val: {chart.series[1].val}")
print(f"Series 1 xVal: {chart.series[1].xVal}")

# Save
chart_sheet = wb.create_chartsheet(title='TestScatter')
chart_sheet.add_chart(chart)
wb.save('test_scatter_correct.xlsx')
print("Saved to test_scatter_correct.xlsx")
