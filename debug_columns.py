from openpyxl import load_workbook
from pathlib import Path

# Load the Data sheet from the Excel file to see the columns
excel_file = Path("C:/Users/EddyS/AppData/Local/Temp/campbell_sci/bm_2011_01_18_10_30_09.xlsx")

wb = load_workbook(excel_file, data_only=False)
ws = wb['Data']

print("Column headers:")
for col_idx in range(1, min(40, ws.max_column + 1)):
    cell = ws.cell(row=1, column=col_idx)
    print(f"  Column {chr(64 + (col_idx % 26) if col_idx <= 26 else 64 + 26 + ((col_idx-1) // 26)) if col_idx <= 26 else chr(64 + ((col_idx-1) // 26)) + chr(65 + ((col_idx-1) % 26))}: {cell.value}")

# Check if we can see the Wind columns
print("\nSearching for 'Wind' columns:")
for col_idx in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col_idx)
    if cell.value and 'Wind' in str(cell.value):
        from openpyxl.utils import get_column_letter
        print(f"  {get_column_letter(col_idx)}: {cell.value}")
