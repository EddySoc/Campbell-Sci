from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference
from openpyxl.chart.data_source import AxDataSource, NumRef
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active

# Simulating: X = Wind_40_m (new), Y = Wind_40_m_old_Avg (old)
ws['A1'] = 'Wind_40_m'      # X-axis
ws['B1'] = 'Wind_40_m_old'  # Y-axis

data = [
    (2.1, 2.0),   # nearly perfect
    (3.5, 3.4),
    (4.2, 4.8),   # drifted up
    (2.8, 2.7),
    (5.0, 5.2),   # drifted up
]

for i, (x, y) in enumerate(data, start=2):
    ws[f'A{i}'] = x
    ws[f'B{i}'] = y

# Create scatter chart the CORRECT way
chart = ScatterChart()
chart.title = "Sensor Comparison"
chart.x_axis.title = "Wind 40m (new)"
chart.y_axis.title = "Wind 40m old"

# For scatter: Y values = column B (including header), X values = column A (data only, no header)
ref_y = Reference(ws, min_col=2, min_row=1, max_row=len(data)+1)  # B1:B6
ref_x = Reference(ws, min_col=1, min_row=2, max_row=len(data)+1)  # A2:A6 (no header!)

# Add Y data (this creates series[0].val)
chart.add_data(ref_y, titles_from_data=True)

# Now set X values - convert reference to AxDataSource
col_letter = get_column_letter(ref_x.min_col)
sheetname = ref_x.sheetname if ref_x.sheetname.startswith("'") else f"'{ref_x.sheetname}'"
formula = f"{sheetname}!${col_letter}${ref_x.min_row}:${col_letter}${ref_x.max_row}"
numref = NumRef(f=formula)
chart.series[0].xVal = AxDataSource(numRef=numref)

print(f"Series 0 val: {chart.series[0].val}")
print(f"Series 0 xVal: {chart.series[0].xVal}")

# Save
chart_sheet = wb.create_chartsheet(title='Scatter')
chart_sheet.add_chart(chart)
wb.save('test_scatter_output.xlsx')
print("Saved to test_scatter_output.xlsx")

