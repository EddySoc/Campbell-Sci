from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.data_source import AxDataSource, NumRef
from openpyxl.utils import get_column_letter

def _reference_formula(ref: Reference) -> str:
    """Build a formula string from a Reference object for use with NumRef."""
    col_letter = get_column_letter(ref.min_col)
    sheetname = ref.sheetname if ref.sheetname.startswith("'") else f"'{ref.sheetname}'"
    return f"{sheetname}!${col_letter}${ref.min_row}:${col_letter}${ref.max_row}"

wb = Workbook()
ws = wb.active

# Add data WITHOUT header in data column
ws['A1'] = 'X'
ws['B1'] = 'Y1'
for i in range(2, 10):
    ws[f'A{i}'] = i
    ws[f'B{i}'] = i * 2

# Create scatter chart
chart = ScatterChart()

# Create Reference objects - WITHOUT header for scatter
cat_idx = 1
data_col = 2
first_data_row = 2
last_row = 9

categories = Reference(ws, min_col=cat_idx, min_row=first_data_row, max_row=last_row)
data = Reference(ws, min_col=data_col, min_row=first_data_row, max_row=last_row)  # NO HEADER

print("Using add_data() then overriding xVal...")

# Add data to create val
chart.add_data(data, titles_from_data=False)
print(f"After add_data: Series[0].val = {chart.series[0].val}")

# Override xVal
xval_formula = _reference_formula(categories)
chart.series[0].xVal = AxDataSource(numRef=NumRef(f=xval_formula))
print(f"After setting xVal: Series[0].xVal = {chart.series[0].xVal}")

# Set title
chart.series[0].title = SeriesLabel(v="TestData")

# Create chart sheet and save
chart_sheet = wb.create_chartsheet(title='TestChart')
chart_sheet.add_chart(chart)
wb.save('test_add_data_then_xval.xlsx')

print("Saved to test_add_data_then_xval.xlsx")

# Verify XML
import zipfile
with zipfile.ZipFile('test_add_data_then_xval.xlsx', 'r') as z:
    content = z.read('xl/charts/chart1.xml').decode('utf-8')
    if '<ser>' in content:
        start = content.find('<ser>')
        end = content.find('</ser>') + len('</ser>')
        ser_xml = content[start:end]
        
        print("\nGenerated XML:")
        print(f"Has val: {'<val>' in ser_xml}")
        print(f"Has xVal: {'<xVal>' in ser_xml}")
        
        if '<val>' in ser_xml:
            val_start = ser_xml.find('<val>')
            val_end = ser_xml.find('</val>') + len('</val>')
            print(f"val element found!")
        else:
            print("val element MISSING!")
        
        if '<xVal>' in ser_xml:
            xval_start = ser_xml.find('<xVal>')
            xval_end = ser_xml.find('</xVal>') + len('</xVal>')
            print(f"xVal: {ser_xml[xval_start:xval_end][:100]}")
